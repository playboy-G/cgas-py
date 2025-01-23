from types import SimpleNamespace

import openpyxl
from openpyxl.styles import Font


class OperationXlsx:
    def __init__(self, file_path=None, sheet_name=None):
        if file_path is not None:
            self.file_path = file_path
        else:
            return None
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=False)

        if sheet_name is not None:
            self.sheet_name = sheet_name
        else:
            self.sheet_name = self.workbook.get_sheet_names()[0]  # 默认获取第一个sheet的数据

        self.worksheet = self.get_data(self.sheet_name)

    def get_data(self, sheet_name):
        # 通过sheet_name获取Excel的表
        worksheet = self.workbook[sheet_name]
        # tables=worksheet .sheets()[self.sheet_id]
        return worksheet

    # 获取指定sheet获取数据行数
    def get_sheet_lines(self):
        return self.worksheet.max_row
    # 获取列数
    def get_clos(self):
        return self.worksheet.max_column

    # 读取指定单元格值
    def get_cell_value(self, row, col):
        # row,col从1开始计算
        return self.worksheet.cell(row, col).value

    # 读取每一行的值到对象，以标题为key定义对象属性字段
    def read_excel_to_obj(self, start_row, keys):
        # 创建一个空集合
        data_obj_list = []

        # 遍历工作表中的每一行，start_row：从第几行开始为正式数据
        for row in self.worksheet.iter_rows(min_row=start_row, max_row=self.worksheet.max_row, values_only=True):
            data_obj = SimpleNamespace(**dict(zip(keys, row)))
            data_obj_list.append(data_obj)

        return data_obj_list

    # 根据列号获取某一列的内容，data_start_row--数据起始行，第一行为0
    def get_col_values_by_id(self, data_start_row=0, col_id=None):
        values = []
        for cell in list(self.worksheet.columns)[col_id]:
            values.append(cell.value)
        return values[data_start_row:]

    # 根据行号，找到该行的内容
    def get_row_values_by_id(self, row):
        str = []
        for cell in list(self.worksheet.rows)[row]:
            str.append(cell.value)
        return str

    # 根据首列对应的值，找到对应的行号
    def get_row_num(self, case_id):
        num = 0
        # 根据Excel的格式第0列为测试用例编号
        cols_data = self.get_col_values_by_id(0)
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num = num + 1
        return 0

    # 根据首列的值，找到行内容（即单条测试用例的数据）
    def get_row_value_by_string(self, row_value):
        row_num = self.get_row_num(case_id=row_value)
        # 通过行号，获取行的值
        rows_data = self.get_row_values_by_id(row_num)
        return rows_data

    # 根据（行，列）写入数据,测试结果写入，行列是从1开始的,可设置样式
    def write_cell_value(self, row, col, valus):
        """
        Font 对象参数说明：
        Font(
            name=None,      # 字体名，可以用字体名字的字符串
            strike=None,    # 删除线，True/False
            color=None,     # 文字颜色
            size=None,      # 字号
            bold=None,      # 加粗, True/False
            italic=None,    # 倾斜，Tue/False
            underline=None # 下划线, 'singleAccounting', 'double', 'single', 'doubleAccounting'
        )
        """
        # 初始化字体对象
        font_ = Font(
            size=14,
            italic=True,
            color='ff0000',
            bold=False,
            strike=None
        )
        self.worksheet.cell(row, col).font = font_
        self.worksheet.cell(row, col, valus)

        # self.worksheet["B1"]="test"  #sheet值赋值
        self.workbook.save(self.file_path)

    def __del__(self):
        self.workbook.close()

# if __name__ == '__main__':
#     excel = OperationXlsx("test.xlsx", "兼职信息")
#     data_dict = excel.read_excel_to_dict(3)
#     print(data_dict)

